import re
import openpyxl
from openpyxl.utils import column_index_from_string
import PySimpleGUI as sg
import sys
import os

def keypress(event):
    """Обработка копи паста на русской раскладке"""
    if event.keycode == 86 and event.keysym.lower() != 'v':
        event.widget.event_generate('<<Paste>>')
    elif event.keycode == 67 and event.keysym.lower() != 'c':
        event.widget.event_generate('<<Copy>>')
    elif event.keycode == 88 and event.keysym.lower() != 'x':
        event.widget.event_generate('<<Cut>>')

def is_valid_mac(mac):
    """Валидация MAC-адреса"""
    mac = mac.strip().upper()
    mac = re.sub(r'[:\-.]', ':', mac)
    if re.fullmatch(r'^([0-9A-F]{2}:){5}[0-9A-F]{2}$', mac):
        return mac
    return None

def is_valid_uin(uin):
    """Валидация УИН"""
    return re.fullmatch(r'^[A-Z]{2}\d{4}-\d{2}-\d{4}-\d{3}$', uin.strip()) is not None

def get_template_path():
    """Возвращает путь к файлу-шаблону, который может быть встроен в exe."""
    if getattr(sys, 'frozen', False):
        # Запущено как скомпилированный .exe
        base_path = sys._MEIPASS
    else:
        # Запущено как обычный скрипт
        base_path = os.path.dirname(os.path.abspath(__file__))

    template_name = 'Реестр_Строителей.xlsx'
    return os.path.join(base_path, template_name)

# --- GUI ---
sg.theme('Default1')

layout = [
    [sg.Text('📁 Как назовем файл:'), sg.Input('Реестр_Строителей_заполненный.xlsx', key='-OUTPUT-', size=(50, 1)),
     sg.FileSaveAs(file_types=(("Excel files", "*.xlsx"),))],

    [sg.HorizontalSeparator()],
    [sg.Text('ОБЩИЕ ДАННЫЕ ПО ОБЪЕКТУ', font=('Arial', 12, 'bold'))],

    [sg.Text('Наименование объекта', size=(25, 1)), sg.Input(
        ' Жилой дом с инженерными сетями и благоустройством', key='-OBJECT_NAME-', size=(70, 1))],
    [sg.Text('Адрес объекта', size=(25, 1)), sg.Input(key='-OBJECT_ADDRESS-', size=(70, 1))],
    [sg.Text('УИН объекта', size=(25, 1)), sg.Input(key='-UIN-', size=(40, 1))],
    [sg.Text('ИНН генподрядчика', size=(25, 1)), sg.Input(key='-INN_GEN-', size=(40, 1))],
    [sg.Text('КПП генподрядчика', size=(25, 1)), sg.Input(key='-KPP_GEN-', size=(40, 1))],
    [sg.Text('Дата начала', size=(25, 1)), sg.Input('1', key='-DATE_START-', size=(40, 1))],
    [sg.Text('Дата окончания', size=(25, 1)), sg.Input('1', key='-DATE_END-', size=(40, 1))],
    [sg.Text('Тип прибора', size=(25, 1)), sg.Input('СКУД', key='-DEVICE_TYPE-', size=(40, 1))],
    [sg.Text('Дата ввода', size=(25, 1)), sg.Input(key='-DATE_ACTIVATE-', size=(40, 1))],
    [sg.Text('Дата снятия', size=(25, 1)), sg.Input('1', key='-DATE_REMOVE-', size=(40, 1))],
    [sg.Text('Даты простоя', size=(25, 1)), sg.Input('1', key='-DOWNTIME-', size=(40, 1))],

    [sg.HorizontalSeparator()],
    [sg.Text('📟 КОЛИЧЕСТВО ПРИБОРОВ', font=('Arial', 12, 'bold'))],
    [sg.Text('Укажите количество приборов', size=(25, 2)), sg.Input('1', key='-COUNT-', size=(10, 1)),
     ],

    [sg.HorizontalSeparator()],
    [sg.Text('📋 СПИСОК ПРИБОРОВ', font=('Arial', 12, 'bold'))],
    [sg.Column([
        [sg.Text('№', size=(4, 1)), sg.Text('Наименование прибора', size=(20, 1)), sg.Text('MAC-адрес', size=(25, 1))],
        [sg.Listbox(values=[], key='-DEVICE_LIST-', size=(70, 8), select_mode=sg.LISTBOX_SELECT_MODE_SINGLE)]
    ])],
    [sg.Text('Наименование', size=(15, 1)), sg.Input('КПП1 вход1', key='-DEVICE_NAME-', size=(30, 1))],
    [sg.Text('MAC-адрес', size=(15, 1)), sg.Input(key='-DEVICE_MAC-', size=(30, 1))],
    [sg.Button('➕ Добавить прибор'), sg.Button('❌ Удалить выбранный'), sg.Button('🗑️ Очистить список')],

    [sg.HorizontalSeparator()],
    [sg.Button('💾 Сохранить в Excel'), sg.Button('Выход')]
]
# sg.Button('Подготовить список приборов')
window = sg.Window('Заполнение реестра строителей', layout, finalize=True)

window.TKroot.bind_all("<Control-KeyPress>", keypress) # Обработка копи паста на русской раскладке

devices = []  # Список кортежей (наименование, mac)

def add_device_to_list():
    name = window['-DEVICE_NAME-'].get().strip()
    mac = window['-DEVICE_MAC-'].get().strip()
    if not name:
        sg.popup_error('Введите наименование прибора!')
        return
    mac_valid = is_valid_mac(mac)
    if not mac_valid:
        sg.popup_error('MAC невалидный. Ожидается XX:XX:XX:XX:XX:XX или XX-XX-XX-XX-XX-XX')
        return
    devices.append((name, mac_valid))
    update_device_list()
    window['-DEVICE_NAME-'].set_focus()


def update_device_list():
    display = [f'{i + 1:2}. {name:30} | {mac}' for i, (name, mac) in enumerate(devices)]
    window['-DEVICE_LIST-'].update(values=display)


while True:
    event, values = window.read()
    if event in (sg.WIN_CLOSED, 'Выход'):
        break

    '''if event == 'Подготовить таблицу приборов':
        try:
            count = int(values['-COUNT-'])
            if count < 1:
                sg.popup_error('Количество должно быть >= 1')
                continue
            #devices.clear()
            update_device_list(count)
            sg.popup_ok(f'Готово! Можно добавлять до {count} приборов (сейчас {len(devices)})')
        except ValueError:
            sg.popup_error('Введите целое число')'''

    if event == '➕ Добавить прибор':
        if devices and len(devices) >= int(values['-COUNT-']):
            sg.popup_error(f'Вы уже добавили {len(devices)} приборов. Задано: {values["-COUNT-"]}')
            continue
        add_device_to_list()

    if event == '❌ Удалить выбранный':
        selected = values['-DEVICE_LIST-']
        if selected:
            idx = int(selected[0].split('.')[0]) - 1
            if 0 <= idx < len(devices):
                del devices[idx]
                update_device_list()

    if event == '🗑️ Очистить список':
        devices.clear()
        update_device_list()

    if event == '💾 Сохранить в Excel':
        uin = values['-UIN-'].strip()

        if not is_valid_uin(uin):
            sg.popup_error(f'УИН "{uin}" невалидный')
            continue
        if not devices:
            sg.popup_error('Добавьте хотя бы один прибор')
            continue

        try:
            expected_count = int(values['-COUNT-'])
            if len(devices) != expected_count:
                sg.popup_error(f'Добавлено {len(devices)} приборов, а задано {expected_count}. Продолжить?')
                response = sg.popup_yes_no('Продолжить запись?')
                if response != 'Yes':
                    continue
        except:
            pass

        template_file = get_template_path()
        try:
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active
        except FileNotFoundError:
            sg.popup_error(f'Файл шаблона "{template_file}" не найден!')
            continue
        except Exception as e:
            sg.popup_error(f'Ошибка открытия шаблона: {e}')
            continue

        common = {
            'D': values['-OBJECT_NAME-'].strip(),
            'E': values['-UIN-'].strip(),
            'F': values['-OBJECT_ADDRESS-'].strip(),
            'G': values['-INN_GEN-'].strip(),
            'H': values['-KPP_GEN-'].strip(),
            'I': values['-DATE_START-'].strip(),
            'J': values['-DATE_END-'].strip(),
            'N': values['-DEVICE_TYPE-'].strip(),
            'P': values['-DATE_ACTIVATE-'].strip(),
            'Q': values['-DATE_REMOVE-'].strip(),
            'R': values['-DOWNTIME-'].strip()
        }

        for i, (device_name, mac) in enumerate(devices):
            row = 2 + i
            # Колонки A, B, C не трогаем – оставляем как в шаблоне
            ws.cell(row=row, column=column_index_from_string('D')).value = common['D']
            ws.cell(row=row, column=column_index_from_string('E')).value = common['E']
            ws.cell(row=row, column=column_index_from_string('F')).value = common['F']
            ws.cell(row=row, column=column_index_from_string('G')).value = common['G']
            ws.cell(row=row, column=column_index_from_string('H')).value = common['H']
            ws.cell(row=row, column=column_index_from_string('I')).value = common['I']
            ws.cell(row=row, column=column_index_from_string('J')).value = common['J']
            ws.cell(row=row, column=column_index_from_string('N')).value = common['N']
            ws.cell(row=row, column=column_index_from_string('P')).value = common['P']
            ws.cell(row=row, column=column_index_from_string('Q')).value = common['Q']
            ws.cell(row=row, column=column_index_from_string('R')).value = common['R']

            device_id = f'{uin}-{mac}'
            ws.cell(row=row, column=column_index_from_string('K')).value = device_id
            ws.cell(row=row, column=column_index_from_string('L')).value = mac
            ws.cell(row=row, column=column_index_from_string('M')).value = device_name
            ws.cell(row=row, column=column_index_from_string('O')).value = mac

        output_file = values['-OUTPUT-'].strip()
        if not output_file:
            output_file = 'Реестр_Строителей_заполненный.xlsx'
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        try:
            wb.save(output_file)
            sg.popup_ok(f'✅ Успешно сохранено!\nФайл: {output_file}\nЗаполнено строк: {len(devices)}')
        except Exception as e:
            sg.popup_error(f'Ошибка сохранения: {e}')